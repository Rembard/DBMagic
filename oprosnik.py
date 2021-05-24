from clickhouse_driver import Client
import openpyxl
from openpyxl import Workbook, load_workbook 
from openpyxl.styles import Alignment, Font, Side, Border, PatternFill
from exchangelib import Credentials
from datetime import datetime, date, timedelta
import numpy as np
import time
import os
import pyodbc
import pymysql
import warnings
import csv
import re

#Старт работы приложения
startAll = time.time()
print('='*40)

path = "C:\\Users\\artur.rakhmanov\\Projects\\DBMagic\\Опросник.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
max_col = sheet_obj.max_column
max_row = sheet_obj.max_row
data2 = []
for n in range(5, max_row+1):
    data = []
    answer6 = ''
    answer7 = ''
    answer8 = ''
    answer9 = ''
    answer10 = ''
    answer11 = ''
    answer12 = ''
    answer14 = ''
    answer15 = ''
    answer16 = ''
    answer17 = ''
    answer18 = ''
    answer19 = ''
    answer20 = ''
    answer21 = ''
    answer22 = ''
    an221 = ''
    an222 = ''
    an223 = ''
    an224 = ''
    an225 = ''
    an226 = ''
    an227 = ''
    an228 = ''
    answer23 = ''
    answer24 = ''
    answer25 = ''
    answer26 = ''
    answer27 = ''
    answer28 = ''
    answer29 = ''
    answer30 = ''
    answer31 = ''
    answer32 = ''
    answer33 = ''
    answer34 = ''
    answer35 = ''
    answer36 = ''
    answer37 = ''
    answer38 = ''
    answer39 = ''
    answer40 = ''
    answer41 = ''
    answer42 = ''
    answer43 = ''
    for i in range(1, max_col + 1):
        dataset = sheet_obj.cell(row = n, column = i)
        if (i == 1 or i == 9 or i == 10 or i ==11 or i ==12 or i ==13):
            data.append(dataset.value)
        if any(x == i for x in range(14,34)):
            if dataset.value != None:
                quetion6 = sheet_obj.cell(row = 3, column = i)
                # print (quetion6.value)
                if answer6 != '':
                    answer6 = answer6+ '\n' + quetion6.value
                else:
                    answer6 = quetion6.value
        if any(x == i for x in range(34,62)):
            if dataset.value != None:
                quetion7 = sheet_obj.cell(row = 3, column = i)
                if answer7 != '':
                    answer7 = answer7+ '\n' + quetion7.value
                else:
                    answer7 = quetion7.value
        if any(x == i for x in range(62,70)):
            if dataset.value != None:
                quetion8 = sheet_obj.cell(row = 3, column = i)
                if answer8 != '':
                    answer8 = answer8+ '\n' + quetion8.value
                else:
                    answer8 = quetion8.value
        if any(x == i for x in range(70,72)):
            if dataset.value != None:
                quetion9 = sheet_obj.cell(row = 3, column = i)
                if answer9 != '':
                    answer9 = answer9+ '\n' + quetion9.value
                else:
                    answer9 = quetion9.value
        if i == 72:
            if dataset.value != None:
                answer10 = dataset.value
        if any(x == i for x in range(73,93)):
            if dataset.value != None:
                quetion11 = sheet_obj.cell(row = 3, column = i)
                if answer11 != '':
                    answer11 = answer11+ '\n' + quetion11.value
                else:
                    answer11 = quetion11.value
        if any(x == i for x in range(93,121)):
            if dataset.value != None:
                quetion12 = sheet_obj.cell(row = 3, column = i)
                if answer12 != '':
                    answer12 = answer12+ '\n' + quetion12.value
                else:
                    answer12 = quetion12.value
        if any(x == i for x in range(121,123)):
            if dataset.value != None:
                quetion14 = sheet_obj.cell(row = 3, column = i)
                if answer14 != '':
                    answer14 = answer14+ '\n' + quetion14.value
                else:
                    answer14 = quetion14.value
        if i == 123:
            if dataset.value != None:
                answer15 = dataset.value
        if any(x == i for x in range(124,144)):
            if dataset.value != None:
                quetion16 = sheet_obj.cell(row = 3, column = i)
                if answer16 != '':
                    answer16 = answer16+ '\n' + quetion16.value
                else:
                    answer16 = quetion16.value
        if any(x == i for x in range(144,172)):
            if dataset.value != None:
                quetion17 = sheet_obj.cell(row = 3, column = i)
                if answer17 != '':
                    answer17 = answer17+ '\n' + quetion17.value
                else:
                    answer17 = quetion17.value
        if any(x == i for x in range(172,180)):
            if dataset.value != None:
                quetion18 = sheet_obj.cell(row = 3, column = i)
                if answer18 != '':
                    answer18 = answer18+ '\n' + quetion18.value
                else:
                    answer18 = quetion18.value
        if any(x == i for x in range(180,182)):
            if dataset.value != None:
                quetion19 = sheet_obj.cell(row = 3, column = i)
                if answer19 != '':
                    answer19 = answer19+ '\n' + quetion19.value
                else:
                    answer19 = quetion19.value
        if i == 182:
            if dataset.value != None:
                answer20 = dataset.value
        if any(x == i for x in range(183,211)):
            if dataset.value != None:
                quetion21 = sheet_obj.cell(row = 3, column = i)
                if answer21 != '':
                    answer21 = answer21+ '\n' + quetion21.value
                else:
                    answer21 = quetion21.value
        if any(x == i for x in range(211,219)):
            if dataset.value != None:
                quetion22 = sheet_obj.cell(row = 3, column = i)
                if dataset.value == 1:
                    an221 = '1 - '+ quetion22.value
                if dataset.value == 2:
                    an222 = '2 - '+ quetion22.value
                if dataset.value == 3:
                    an223 = '3 - '+ quetion22.value
                if dataset.value == 4:
                    an224 = '4 - '+ quetion22.value
                if dataset.value == 5:
                    an225 = '5 - '+ quetion22.value
                if dataset.value == 6:
                    an226 = '6 - '+ quetion22.value
                if dataset.value == 7:
                    an227 = '7 - '+ quetion22.value
                if dataset.value == 8:
                    an228 = '8 - '+ quetion22.value
                answer22 = an221+'\n'+an222+'\n'+an223+'\n'+an224+'\n'+an225+'\n'+an226+'\n'+an227+'\n'+an228
        if any(x == i for x in range(219,236)):
            if dataset.value != None:
                quetion23 = sheet_obj.cell(row = 3, column = i)
                if answer23 != '':
                    answer23 = answer23+ '\n' + quetion23.value
                else:
                    answer23 = quetion23.value
        if any(x == i for x in range(236,323)):
            if dataset.value != None:
                quetion24 = sheet_obj.cell(row = 3, column = i)
                if answer24 != '':
                    answer24 = answer24+ '\n' + quetion24.value
                else:
                    answer24 = quetion24.value
        if any(x == i for x in range(323,377)):
            if dataset.value != None:
                quetion25 = sheet_obj.cell(row = 3, column = i)
                if answer25 != '':
                    answer25 = answer25+ '\n' + quetion25.value
                else:
                    answer25 = quetion25.value
        if any(x == i for x in range(377,424)):
            if dataset.value != None:
                quetion26 = sheet_obj.cell(row = 3, column = i)
                if answer26 != '':
                    answer26 = answer26+ '\n' + quetion26.value
                else:
                    answer26 = quetion26.value
        if any(x == i for x in range(424,456)):
            if dataset.value != None:
                quetion27 = sheet_obj.cell(row = 3, column = i)
                if answer27 != '':
                    answer27 = answer27+ '\n' + quetion27.value
                else:
                    answer27 = quetion27.value
        if any(x == i for x in range(456,480)):
            if dataset.value != None:
                quetion28 = sheet_obj.cell(row = 3, column = i)
                if answer28 != '':
                    answer28 = answer28+ '\n' + quetion28.value
                else:
                    answer28 = quetion28.value
        if any(x == i for x in range(480,525)):
            if dataset.value != None:
                quetion29 = sheet_obj.cell(row = 3, column = i)
                if answer29 != '':
                    answer29 = answer29+ '\n' + quetion28.value
                else:
                    answer29 = quetion29.value
        if any(x == i for x in range(525,601)):
            if dataset.value != None:
                quetion30 = sheet_obj.cell(row = 3, column = i)
                if answer30 != '':
                    answer30 = answer30+ '\n' + quetion30.value
                else:
                    answer30 = quetion30.value
        if any(x == i for x in range(601,638)):
                if dataset.value != None:
                    quetion31 = sheet_obj.cell(row = 3, column = i)
                    if answer31 != '':
                        answer31 = answer31+ '\n' + quetion31.value
                    else:
                        answer31 = quetion31.value
        if any(x == i for x in range(638,668)):
                if dataset.value != None:
                    quetion32 = sheet_obj.cell(row = 3, column = i)
                    if answer32 != '':
                        answer32 = answer32+ '\n' + quetion32.value
                    else:
                        answer32 = quetion32.value
        if any(x == i for x in range(668,713)):
                if dataset.value != None:
                    quetion33 = sheet_obj.cell(row = 3, column = i)
                    if answer33 != '':
                        answer33 = answer33+ '\n' + quetion33.value
                    else:
                        answer33 = quetion33.value
        if any(x == i for x in range(713,743)):
                if dataset.value != None:
                    quetion34 = sheet_obj.cell(row = 3, column = i)
                    if answer34 != '':
                        answer34 = answer34+ '\n' + quetion34.value
                    else:
                        answer34 = quetion33.value
        if any(x == i for x in range(743,773)):
                if dataset.value != None:
                    quetion35 = sheet_obj.cell(row = 3, column = i)
                    if answer35 != '':
                        answer35 = answer35+ '\n' + quetion35.value
                    else:
                        answer35 = quetion35.value
        if any(x == i for x in range(773,827)):
                if dataset.value != None:
                    quetion36 = sheet_obj.cell(row = 3, column = i)
                    if answer36 != '':
                        answer36 = answer36+ '\n' + quetion36.value
                    else:
                        answer36 = quetion36.value
        if any(x == i for x in range(827,854)):
                quetion37 = sheet_obj.cell(row = 3, column = i)
                print (quetion37.value)
                if dataset.value != None:
                    quetion37 = sheet_obj.cell(row = 3, column = i)
                    if answer37 != '':
                        answer37 = answer37+ '\n' + quetion37.value
                    else:
                        answer37 = quetion37.value
        if any(x == i for x in range(854,880)):
                quetion38 = sheet_obj.cell(row = 3, column = i)
                print (quetion38.value)
                if dataset.value != None:
                    quetion38 = sheet_obj.cell(row = 3, column = i)
                    if answer38 != '':
                        answer38 = answer38+ '\n' + quetion38.value
                    else:
                        answer38 = quetion38.value
        if i == 880:
            if dataset.value != None:
                answer39 = dataset.value
        if i == 881:
            if dataset.value != None:
                answer40 = dataset.value
        if i == 882:
            if dataset.value != None:
                answer41 = dataset.value
    data.append(answer6)
    data.append(answer7)
    data.append(answer8)
    data.append(answer9)
    data.append(answer10)
    data.append(answer11)
    data.append(answer12)
    data.append(answer14)
    data.append(answer15)
    data.append(answer16)
    data.append(answer17)
    data.append(answer18)
    data.append(answer19)
    data.append(answer20)
    data.append(answer21)
    data.append(answer22)
    data.append(answer23)
    data.append(answer24)
    data.append(answer25)
    data.append(answer26)
    data.append(answer27)
    data.append(answer28)
    data.append(answer29)
    data.append(answer30)
    data.append(answer31)
    data.append(answer32)
    data.append(answer33)
    data.append(answer34)
    data.append(answer35)
    data.append(answer36)
    data.append(answer37)
    data.append(answer38)
    data.append(answer39)
    data.append(answer40)
    data.append(answer41)
    data2.append(data)

BDt = pymysql.connect(host='10.144.198.132',port=3306,user='mon_report_bti',password='Report_112233',db='test',charset="utf8")
BDtest = BDt.cursor()
BDtest.execute("""SET NAMES `utf8`""")
ins = 'INSERT INTO Anketa_test VALUES', data2
rs = str(ins).replace("[","(")
us = rs.replace("('","")
us = us.replace("VALUES',","VALUES")
us = us.replace("((","(")
us = us.replace("]",")")
us = us.replace(")))",")")
BDtest.execute("""DROP TABLE IF EXISTS Anketa_test""")
BDtest.execute("""
        CREATE TABLE IF NOT EXISTS Anketa_test (
        ID int(32), Login varchar(128), FIO varchar(128), Email varchar(128), Number varchar(128),
        Region varchar(128), Question6 varchar(255), Question7 varchar(255), Question8 varchar(255),
        Question9 varchar(255), Question10 varchar(255), Question11 varchar(255), Question12 varchar(255),
        Question14 varchar(255), Question15 varchar(255), Question16 varchar(255), Question17 varchar(255),
        Question18 varchar(255), Question19 varchar(255), Question20 varchar(255), Question21 varchar(512),
        Question22 varchar(512), Question23 varchar(512), Question24 varchar(512), Question25 varchar(512),
        Question26 varchar(512), Question27 varchar(512), Question28 varchar(512), Question29 varchar(512),
        Question30 varchar(512), Question31 varchar(512), Question32 varchar(512), Question33 varchar(512),
        Question34 varchar(512), Question35 varchar(512), Question36 varchar(512), Question37 varchar(512),
        Question38 varchar(512), Question39 varchar(512), Question40 varchar(512), Question41 varchar(512))
        CHARACTER SET utf8 COLLATE utf8_general_ci
        """)
BDtest.execute(us)
BDt.commit()
end = time.time()
print (datetime.now(), "Скрипт работает. Запись произведена. Время исполнения-", end-startAll)
# ID Login FIO Email Number Region Question6[] Question7 Question8 Question9 Question10(str) 
datainsert = []